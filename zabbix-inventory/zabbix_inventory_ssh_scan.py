from __future__ import annotations

import csv
import os
import re
import sys
from dataclasses import dataclass
from pathlib import PurePosixPath
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook

try:
    import inventory_ssh_config as ssh_cfg
except ImportError:
    import inventory_ssh_config_example as ssh_cfg  # placeholder values

try:
    import paramiko
except ImportError as e:
    print("paramiko is required: pip install -r requirements.txt", file=sys.stderr)
    raise


@dataclass
class ServiceInfo:
    host: str
    domain: str
    unit: str
    active_state: Optional[str] = None
    main_pid: Optional[str] = None
    fragment_path: Optional[str] = None
    exec_start: Optional[str] = None
    env_file: Optional[str] = None
    account: Optional[str] = None


@dataclass
class PathInfo:
    host: str
    domain: str
    path: str
    kind: str
    exists: Optional[bool] = None
    obj_type: Optional[str] = None  # file/dir/other
    source: str = "inventory"
    account: Optional[str] = None


def normalize_path(path: str) -> str:
    path = (path or "").strip()
    return path


def classify_path(path: str) -> str:
    p = normalize_path(path)
    low = p.lower()

    if any(low.endswith(ext) for ext in (".crt", ".cer", ".pem", ".der", ".key", ".p12", ".pfx")):
        return "cert_file"

    # logs
    if any(x in low for x in ("/log/", "/logs/")) or low.endswith(".log") or low.endswith(".log.1") or ".log." in low:
        return "log_file"

    # configs
    if low.startswith("/etc"):
        return "config_file"
    if any(low.endswith(ext) for ext in (".conf", ".ini", ".cfg", ".yaml", ".yml", ".json", ".xml", ".properties")):
        return "config_file"
    if any(low.endswith(ext) for ext in (".service", ".target", ".socket", ".timer", ".mount", ".slice")):
        return "config_file"

    # binaries
    if low.startswith("/bin/") or low.startswith("/sbin/") or low.startswith("/usr/bin/") or low.startswith("/usr/sbin/"):
        return "binary_file"
    if low.startswith("/usr/local/bin/") or low.startswith("/usr/local/sbin/"):
        return "binary_file"
    if any(low.endswith(ext) for ext in (".sh", ".py", ".pl", ".rb")):
        return "binary_file"

    if low.startswith("/var/lib") or low.startswith("/var/opt") or low.startswith("/var/www") or low.startswith("/srv") or low.startswith("/home") or low.startswith("/opt"):
        return "data_file"

    return "other"


def derive_dir_kind_from_file_kind(file_kind: str) -> str:
    if file_kind == "log_file":
        return "log_dir"
    if file_kind == "config_file":
        return "config_dir"
    if file_kind == "cert_file":
        return "cert_dir"
    if file_kind == "binary_file":
        return "binary_dir"
    if file_kind == "data_file":
        return "data_dir"
    return "other"


def find_domain(host: str) -> Optional[str]:
    host = (host or "").strip().lower()
    if not host:
        return None
    domains = getattr(ssh_cfg, "DOMAIN_ACCOUNTS", {})
    # longest suffix wins
    matches = [d for d in domains.keys() if host.endswith(d)]
    if not matches:
        return None
    return sorted(matches, key=len, reverse=True)[0]


def parse_services_and_paths(xlsx_path: str) -> tuple[list[ServiceInfo], list[PathInfo]]:
    wb = load_workbook(xlsx_path)

    services: list[ServiceInfo] = []
    paths: list[PathInfo] = []

    if "Services" in wb.sheetnames:
        ws = wb["Services"]
        headers = [str(c.value or "").strip() for c in ws[1]]
        try:
            host_idx = headers.index("Host")
        except ValueError:
            host_idx = None
        try:
            unit_idx = headers.index("Unit")
        except ValueError:
            unit_idx = None
        if host_idx is not None and unit_idx is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                host = (row[host_idx] or "").strip()
                unit = (row[unit_idx] or "").strip()
                if not host or not unit:
                    continue
                domain = find_domain(host) or ""
                account = ssh_cfg.DOMAIN_ACCOUNTS.get(domain, {}).get("username")
                services.append(ServiceInfo(host=host, domain=domain, unit=unit, account=account))

    # Prefer Paths sheet if exists; otherwise extract from categorized sheets
    if "Paths" in wb.sheetnames:
        ws = wb["Paths"]
        headers = [str(c.value or "").strip() for c in ws[1]]
        host_idx = headers.index("Host") if "Host" in headers else None
        kind_idx = headers.index("Kind") if "Kind" in headers else None
        path_idx = headers.index("Path") if "Path" in headers else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if host_idx is None or path_idx is None:
                continue
            host = (row[host_idx] or "").strip()
            if not host:
                continue
            domain = find_domain(host) or ""
            account = ssh_cfg.DOMAIN_ACCOUNTS.get(domain, {}).get("username")
            path = normalize_path(row[path_idx] or "")
            if not path:
                continue
            kind = (row[kind_idx] or "").strip() if kind_idx is not None else classify_path(path)
            paths.append(PathInfo(host=host, domain=domain, path=path, kind=kind, account=account, source="inventory"))
    else:
        # any sheet with columns Host/Path
        for name in wb.sheetnames:
            if name in ("Summary", "Services"):
                continue
            ws = wb[name]
            headers = [str(c.value or "").strip() for c in ws[1]]
            if "Host" not in headers:
                continue
            host_idx = headers.index("Host")
            path_idx = headers.index("Path") if "Path" in headers else None
            if path_idx is None and len(headers) >= 3:
                path_idx = 2  # heuristic
            kind_idx = headers.index("Kind") if "Kind" in headers else None

            for row in ws.iter_rows(min_row=2, values_only=True):
                host = (row[host_idx] or "").strip()
                if not host:
                    continue
                domain = find_domain(host) or ""
                account = ssh_cfg.DOMAIN_ACCOUNTS.get(domain, {}).get("username")
                path = normalize_path(row[path_idx] or "") if path_idx is not None else ""
                if not path:
                    continue
                kind = (row[kind_idx] or "").strip() if kind_idx is not None else classify_path(path)
                paths.append(PathInfo(host=host, domain=domain, path=path, kind=kind, account=account, source=name))

    return services, paths


def parse_systemd_show(output: str) -> Dict[str, str]:
    info: dict[str, str] = {}
    for line in output.splitlines():
        line = line.strip()
        if not line or "=" not in line:
            continue
        k, v = line.split("=", 1)
        info[k] = v
    return info


def extract_exec_binary(exec_start: str) -> Optional[str]:
    # ExecStart may look like: /usr/bin/python3 -m module
    if not exec_start:
        return None
    # systemd show: ExecStart={ path=..., argv[]=...; } can be messy
    # but often the first word is the binary
    parts = exec_start.split()
    for w in parts:
        if w.startswith("/"):
            return w
    return None


def ssh_run(client: paramiko.SSHClient, command: str) -> tuple[str, str, int]:
    stdin, stdout, stderr = client.exec_command(command)
    out = stdout.read().decode("utf-8", errors="ignore")
    err = stderr.read().decode("utf-8", errors="ignore")
    rc = stdout.channel.recv_exit_status()
    return out, err, rc


def ssh_client_for(host: str, domain: Optional[str]) -> paramiko.SSHClient:
    cfg: dict[str, Any] = {}
    if domain:
        cfg = ssh_cfg.DOMAIN_ACCOUNTS.get(domain, {})
    username = cfg.get("username") or None
    password = cfg.get("password") or None
    key_filename = cfg.get("key_filename") or None
    port = getattr(ssh_cfg, "DEFAULT_SSH_PORT", 22)
    timeout = getattr(ssh_cfg, "SSH_CONNECT_TIMEOUT_SEC", 10)

    c = paramiko.SSHClient()
    c.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    c.connect(hostname=host, port=port, username=username, password=password or None, key_filename=key_filename or None, timeout=timeout)
    return c


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print("Usage: python zabbix_inventory_ssh_scan.py <zabbix_inventory_report.xlsx>", file=sys.stderr)
        return 2

    input_xlsx = argv[1]
    if not os.path.exists(input_xlsx):
        print(f"Input XLSX not found: {input_xlsx}", file=sys.stderr)
        return 1

    services, paths = parse_services_and_paths(input_xlsx)

    # Group by host for SSH sessions
    svc_by_host: dict[str, list[ServiceInfo]] = {}
    for s in services:
        svc_by_host.setdefault(s.host, []).append(s)

    paths_by_host: dict[str, list[PathInfo]] = {}
    for p in paths:
        paths_by_host.setdefault(p.host, []).append(p)

    # Derived paths from systemd units
    for host, svc_list in list(svc_by_host.items()):
        domain = find_domain(host) or ""
        try:
            client = ssh_client_for(host, domain)
        except Exception as e:
            # mark all host services unreachable
            print(f"[ERR] {host}: cannot connect ({e})", file=sys.stderr)
            continue

        try:
            for s in svc_list:
                out, err, rc = ssh_run(client, f"systemctl show -p ExecStart -p FragmentPath -p ActiveState -p MainPID -p EnvironmentFile {s.unit}")
                if rc != 0:
                    continue
                info = parse_systemd_show(out)
                s.active_state = info.get("ActiveState")
                s.main_pid = info.get("MainPID")
                s.fragment_path = info.get("FragmentPath")
                s.exec_start = info.get("ExecStart")
                s.env_file = info.get("EnvironmentFile")

                derived: list[tuple[str, str]] = []

                if s.fragment_path:
                    derived.append((s.fragment_path, "config_file"))

                if s.env_file:
                    for envf in s.env_file.split():
                        derived.append((envf, classify_path(envf)))

                bin_path = extract_exec_binary(s.exec_start or "")
                if bin_path:
                    derived.append((bin_path, "binary_file"))

                    try:
                        bin_dir = str(PurePosixPath(bin_path).parent)
                        derived.append((bin_dir, "service_dir"))
                    except Exception:
                        pass

                # Add derived paths into path list
                host_paths = paths_by_host.setdefault(host, [])
                for pth, kind in derived:
                    pth = normalize_path(pth)
                    if not pth:
                        continue
                    # avoid duplicates
                    key = (s.host, pth)
                    if any(pp.path == pth for pp in host_paths):
                        continue
                    host_paths.append(PathInfo(host=s.host, domain=s.domain, path=pth, kind=kind, source="systemd", account=s.account))
        finally:
            client.close()

    # Now verify paths on hosts
    verified_paths: list[PathInfo] = []
    errors: list[tuple[str, str]] = []

    for host, host_paths in paths_by_host.items():
        domain = find_domain(host) or ""
        try:
            client = ssh_client_for(host, domain)
        except Exception as e:
            errors.append((host, f"connect failed: {e}"))
            continue

        try:
            # Build tab-separated list for bash
            lines: list[str] = []
            for p in host_paths:
                k = p.kind or classify_path(p.path)
                lines.append(f"{p.path}\t{k}\t{p.source}")

            heredoc = "\n".join(lines)
            cmd = (
                "while IFS=$'\t' read -r p kind src; do "
                "if [ -d \"$p\" ]; then type=dir; elif [ -f \"$p\" ]; then type=file; else type=other; fi; "
                "if [ -e \"$p\" ]; then exists=yes; else exists=no; fi; "
                "printf '%s\t%s\t%s\t%s\t%s\n' \"$p\" \"$kind\" \"$exists\" \"$type\" \"$src\"; "
                "done <<'EOF'\n" + heredoc + "\nEOF"
            )
            out, err, rc = ssh_run(client, cmd)
            if rc != 0:
                errors.append((host, f"path scan rc={rc}: {err.strip()}"))

            for line in out.splitlines():
                parts = line.split("\t")
                if len(parts) < 5:
                    continue
                p_path, p_kind, exists, obj_type, src = parts[:5]
                verified_paths.append(
                    PathInfo(
                        host=host,
                        domain=find_domain(host) or "",
                        path=p_path,
                        kind=p_kind or classify_path(p_path),
                        exists=True if exists == "yes" else False,
                        obj_type=obj_type,
                        source=src or "inventory",
                        account=ssh_cfg.DOMAIN_ACCOUNTS.get(domain, {}).get("username"),
                    )
                )
        finally:
            client.close()

    # Build unique dirs
    dirs_set: set[tuple[str, str, str]] = set()
    for p in verified_paths:
        if not p.exists:
            continue
        try:
            posix = PurePosixPath(p.path)
        except Exception:
            continue
        dir_path = str(posix) if p.obj_type == "dir" else str(posix.parent)
        dir_kind = derive_dir_kind_from_file_kind(p.kind)
        dirs_set.add((p.host, dir_path, dir_kind))

    output_xlsx = os.path.splitext(os.path.basename(input_xlsx))[0] + "_ssh_scan.xlsx"

    wb_out = Workbook()

    ws_paths = wb_out.active
    ws_paths.title = "Paths"
    ws_paths.append(["Host", "Domain", "Path", "Kind", "Exists", "Type", "Source", "SSH account"])
    for p in verified_paths:
        ws_paths.append([p.host, p.domain, p.path, p.kind, "yes" if p.exists else "no", p.obj_type, p.source, p.account])

    ws_dirs = wb_out.create_sheet("Dirs")
    ws_dirs.append(["Host", "Domain", "Dir", "Kind", "SSH account"])
    for host, dir_path, kind in sorted(dirs_set):
        domain = find_domain(host) or ""
        account = ssh_cfg.DOMAIN_ACCOUNTS.get(domain, {}).get("username")
        ws_dirs.append([host, domain, dir_path, kind, account])

    ws_services = wb_out.create_sheet("Services")
    ws_services.append(["Host", "Domain", "Unit", "ActiveState", "MainPID", "ExecStart", "FragmentPath", "EnvironmentFile", "SSH account"])
    for host, svc_list in svc_by_host.items():
        for s in svc_list:
            ws_services.append(
                [
                    s.host,
                    s.domain,
                    s.unit,
                    s.active_state or "",
                    s.main_pid or "",
                    (s.exec_start or "").replace("\n", " "),
                    s.fragment_path or "",
                    s.env_file or "",
                    s.account,
                ]
            )

    if errors:
        ws_err = wb_out.create_sheet("Errors")
        ws_err.append(["Host", "Error"])
        for host, msg in errors:
            ws_err.append([host, msg])

    wb_out.save(output_xlsx)
    print(f"[OK] saved: {output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))

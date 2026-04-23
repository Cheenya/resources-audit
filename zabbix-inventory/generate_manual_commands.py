#!/usr/bin/env python3
from __future__ import annotations

import sys
from collections import defaultdict

from openpyxl import load_workbook


DISABLED_HINTS = {
    "disabled",
    "not monitored",
    "unmonitored",
}


def normalize_host(host: str) -> str:
    return host.strip().lower()


def status_is_disabled(status: str | None) -> bool:
    if status is None:
        return False
    s = str(status).strip().lower()
    for hint in DISABLED_HINTS:
        if hint in s:
            return True
    # common shorthand
    if s in {"0", "false", "no"}:
        return True
    return False


def build_host_status_map(wb) -> dict[str, str]:
    """Best-effort extraction: find host+status columns on any sheet."""
    host_status: dict[str, str] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        headers = [str(c.value or "").strip() for c in ws[1]]
        # host column
        host_idx = None
        for i, h in enumerate(headers):
            if h.lower() == "host":
                host_idx = i
                break
        if host_idx is None:
            continue

        # status column candidates
        status_idx = None
        for i, h in enumerate(headers):
            hl = h.lower()
            if hl in {"status", "monitored", "enabled"}:
                status_idx = i
                break
            if hl in {"available"}:  # not perfect, but often there
                status_idx = i
                break
        if status_idx is None:
            # fuzzy match
            for i, h in enumerate(headers):
                hl = h.lower()
                if any(word in hl for word in ["status", "monitor", "enable", "disable"]):
                    status_idx = i
                    break
        if status_idx is None:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            host = row[host_idx]
            status = row[status_idx]
            if not host:
                continue
            host_status[normalize_host(str(host))] = status

    return host_status


def main(input_xlsx: str, output_path: str | None) -> int:
    wb = load_workbook(input_xlsx)

    if "Services" not in wb.sheetnames:
        print("Sheet 'Services' not found", file=sys.stderr)
        return 1

    host_status = build_host_status_map(wb)

    ws = wb["Services"]
    headers = [str(c.value or "").strip() for c in ws[1]]
    try:
        host_idx = headers.index("Host")
        unit_idx = headers.index("Unit")
    except ValueError:
        print("Sheet 'Services' must contain columns Host and Unit", file=sys.stderr)
        return 1

    by_host: dict[str, list[str]] = defaultdict(list)
    skipped_hosts: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        host = (row[host_idx] or "").strip()
        unit = (row[unit_idx] or "").strip()
        if not host or not unit:
            continue

        nh = normalize_host(host)
        status = host_status.get(nh)
        if status_is_disabled(status):
            skipped_hosts.add(host)
            continue

        by_host[host].append(unit)

    out_lines: list[str] = []
    out_lines.append("# Commands generated from XLSX")
    out_lines.append("# Copy the blocks host-by-host into your terminal.")
    out_lines.append("# Sheet 'Services': Host + Unit")
    if skipped_hosts:
        out_lines.append(
            "# NOTE: skipped hosts (disabled/unmonitored in Zabbix): "
            + ", ".join(sorted(skipped_hosts))
        )
    out_lines.append("")

    for host, units in sorted(by_host.items()):
        out_lines.append(f"echo '==== {host} ===='")

        for unit in sorted(set(units)):
            out_lines.append(f"echo '--- {unit} ---'")
            out_lines.append(f"systemctl show -p ExecStart -p FragmentPath -p EnvironmentFile {unit}")
            out_lines.append(f"systemctl status {unit} --no-pager || true")
            out_lines.append(f"journalctl -u {unit} -n 50 --no-pager || true")
            out_lines.append("")
            out_lines.append(f"UNIT_BASE='{unit}'")
            out_lines.append("UNIT_BASE=${UNIT_BASE%%.service*}")
            out_lines.append('ls -ld /etc/*"$UNIT_BASE"* /etc/"$UNIT_BASE" 2>/dev/null || true')
            out_lines.append('ls -ld /var/log/*"$UNIT_BASE"* /var/log/"$UNIT_BASE" 2>/dev/null || true')
            out_lines.append('ls -ld /var/lib/*"$UNIT_BASE"* /var/lib/"$UNIT_BASE" 2>/dev/null || true')
            out_lines.append('ls -ld /opt/*"$UNIT_BASE"* /opt/"$UNIT_BASE" 2>/dev/null || true')
            out_lines.append(
                'find /etc -maxdepth 4 -type f \\
'
                '  \( -iname "*$UNIT_BASE*" -o -iname "*.conf" -o -iname "*.yaml" -o -iname "*.yml" -o -iname "*.json" \) \\
'
                '  2>/dev/null || true'
            )
            out_lines.append(
                'find / -maxdepth 4 -type f \( -iname "*.crt" -o -iname "*.pem" -o -iname "*.cer" -o -iname "*.key" -o -iname "*.pfx" \) \\
'
                '  2>/dev/null | head -n 30 || true'
            )
            out_lines.append("")
        out_lines.append("")

    if not out_lines:
        out_lines.append("# No hosts/units to generate commands for.")

    text = "\n".join(out_lines)

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(text)
        print(f"[OK] wrote commands to {output_path}")
    else:
        sys.stdout.write(text)

    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(
            "Usage: python generate_manual_commands.py <report.xlsx> [output.sh]",
            file=sys.stderr,
        )
        raise SystemExit(2)

    input_xlsx = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) >= 3 else None
    raise SystemExit(main(input_xlsx, output_path))

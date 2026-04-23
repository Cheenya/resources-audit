#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parse inventory XLSX (from zabbix_inventory_collect.py) and produce a domain-aware report.

Usage:
    python zabbix_inventory_collect_from_xlsx.py <input_xlsx> [output_xlsx]

Input sheets used:
- Services
- Paths

Output workbook contains:
- Dirs (unique directories per host/kind)
- Services (with mapped account)
- Paths (with mapped account and reclassified cert paths)

Note: accounts are matched by host suffix (rosap.com -> rosap, dom.ru -> dom).
"""
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
TOP_ALIGN = Alignment(vertical="top", wrap_text=True)

ACCOUNT_BY_DOMAIN = {
    "rosap.com": "rosap",
    "dom.ru": "dom",
}
CERT_EXTS = (".crt", ".pem", ".key", ".csr", ".pfx", ".p12", ".cer")
CERT_DIR_HINTS = ("ssl", "cert", "pki", "ca", "tls")


def style_worksheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = TOP_ALIGN

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = TOP_ALIGN

    for idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(max_len + 2, 60))


def detect_domain(host_ref: str | None) -> str:
    host_ref = (host_ref or "").lower()
    for suffix in ACCOUNT_BY_DOMAIN:
        if host_ref.endswith(suffix):
            return suffix
    return ""


def reclassify_path_kind(path: str | None, kind: str | None) -> str:
    path = path or ""
    kind = kind or ""
    lower = path.lower()

    if lower.endswith(CERT_EXTS) or any(hint in lower for hint in CERT_DIR_HINTS):
        return "certificate_like"

    return kind


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(
            "Usage: python zabbix_inventory_collect_from_xlsx.py <input_xlsx> [output_xlsx]",
            file=sys.stderr,
        )
        return 2

    input_path = Path(argv[1]).expanduser().resolve()
    output_path = (
        Path(argv[2]).expanduser().resolve()
        if len(argv) >= 3
        else input_path.with_name(input_path.stem + "_domain_report.xlsx")
    )

    wb_in = load_workbook(input_path, read_only=False, data_only=True)

    services: list[list[Any]] = []
    paths: list[list[Any]] = []
    dirs: set[tuple[Any, ...]] = set()

    ws_services = wb_in["Services"]
    for row in ws_services.iter_rows(min_row=2, values_only=True):
        system, role, host_ref, unit, field, last_value, last_clock, item_name, item_key = row
        domain = detect_domain(host_ref)
        account = ACCOUNT_BY_DOMAIN.get(domain, "")
        services.append([
            domain,
            account,
            host_ref,
            system,
            role,
            unit,
            field,
            last_value,
            last_clock,
            item_name,
            item_key,
        ])

    ws_paths = wb_in["Paths"]
    for row in ws_paths.iter_rows(min_row=2, values_only=True):
        system, role, host_ref, path, kind, lastvalue_excerpt, last_clock, item_name, item_key = row
        domain = detect_domain(host_ref)
        account = ACCOUNT_BY_DOMAIN.get(domain, "")
        kind = reclassify_path_kind(path, kind)
        paths.append(
            [
                domain,
                account,
                host_ref,
                system,
                role,
                path,
                kind,
                lastvalue_excerpt,
                last_clock,
                item_name,
                item_key,
            ]
        )

        dirname = ""
        try:
            dirname = str(Path(path).expanduser().resolve().parent)
        except Exception:
            pass

        if dirname:
            dirs.add((domain, account, host_ref, system, role, kind, dirname))

    wb_out = Workbook()
    wb_out.active.title = "Dirs"
    ws_out = wb_out.active
    ws_out.append(["Domain", "Account", "Host", "System", "Role", "Kind", "Directory"])
    for d in sorted(dirs):
        ws_out.append(list(d))
    style_worksheet(ws_out)

    def add_sheet(title: str, headers: list[str], rows: list[list[Any]]) -> None:
        ws = wb_out.create_sheet(title)
        ws.append(headers)
        for r in rows:
            ws.append(r)
        style_worksheet(ws)

    add_sheet(
        "Services",
        [
            "Domain",
            "Account",
            "Host",
            "System",
            "Role",
            "Unit",
            "Field",
            "Last Value",
            "Last Clock",
            "Item Name",
            "Item Key",
        ],
        services,
    )

    add_sheet(
        "Paths",
        [
            "Domain",
            "Account",
            "Host",
            "System",
            "Role",
            "Path",
            "Kind",
            "Last Value Excerpt",
            "Last Clock",
            "Item Name",
            "Item Key",
        ],
        paths,
    )

    wb_out.save(output_path)
    print(f"[OK] Report written to: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))

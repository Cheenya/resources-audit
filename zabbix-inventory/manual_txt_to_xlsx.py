#!/usr/bin/env python3
from __future__ import annotations

import os
import re
import sys
from glob import glob
from pathlib import Path
from shlex import split as shlex_split
from typing import Any

from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Manual TXT -> XLSX post-processing
#
# Workflow this script supports:
#  1) Generate a bundle of commands from your inventory XLSX
#  2) SSH to each host, run the commands, save output to /tmp/manual_inventory_<host>.txt
#  3) Copy those *.txt files to a local directory
#  4) Run this script to merge them into one Excel report with full paths
#
# Usage examples:
#   python manual_txt_to_xlsx.py /tmp/manual_inventory_txt/ manual_inventory_report.xlsx
#   python manual_txt_to_xlsx.py manual_inventory_host1.txt manual_inventory_host2.txt
#
# XLSX schema (sheets): Paths, Dirs, Services, Errors
# ---------------------------------------------------------------------------

CERT_EXT = {".crt", ".pem", ".cer", ".key", ".pfx"}


def _is_abs_path_token(token: str) -> bool:
    return token.startswith("/") and not token.startswith("//")


def _guess_host_from_filename(filename: str) -> str:
    base = Path(filename).name
    # inventory_host1.txt -> host1
    for prefix in ("manual_inventory_", "inventory_", "manual_", "out_", "report_"):
        if base.startswith(prefix):
            base = base[len(prefix):]
            break
    base = re.sub(r"\.txt$", "", base)
    return base


def _extract_abs_paths_from_line(line: str) -> list[str]:
    # extremely permissive: all absolute paths and/or space-separated command arguments
    line = line.strip()
    if not line:
        return []

    # ExecStart= can contain quotes; easiest is shlex split after removing key prefix
    if line.startswith("ExecStart="):
        value = line.split("=", 1)[1]
        tokens = shlex_split(value)
        return [t for t in tokens if _is_abs_path_token(t)]

    # systemctl show keys
    if any(line.startswith(p) for p in ("FragmentPath=", "EnvironmentFile=", "DropInPaths=")):
        value = line.split("=", 1)[1]
        return [t for t in re.split(r"\\s+", value) if _is_abs_path_token(t)]

    # find output: each line is a single absolute path
    if line.startswith("/"):
        # avoid capturing something like "/.journal" as fake marker
        # We accept anything that looks like a path component; best effort.
        return [line.split()[0]]

    return []


def classify_path(path: str) -> str:
    p = Path(path)
    # certificates first so /etc/ssl/certs/*.crt doesn't become "config"
    if p.suffix.lower() in CERT_EXT or any(part.lower() in {"ssl", "pki", "certs", "cert"} for part in p.parts):
        return "certificate"
    if "var/log" in p.as_posix():
        return "log"
    if p.suffix.lower() in {".log", ".gz", ".journal"}:
        return "log"
    if p.as_posix().startswith("/etc"):
        return "config"
    if any(p.as_posix().startswith(prefix) for prefix in ("/var/lib", "/srv", "/data")):
        return "data"
    if any(p.as_posix().startswith(prefix) for prefix in ("/usr/bin", "/usr/sbin", "/bin", "/sbin", "/opt")):
        return "binary"
    return "other"


def parse_service_block(unit: str, lines: list[str]) -> dict[str, Any]:
    """Extract ExecStart, fragment, env files and basic status hints from a unit block."""
    exec_paths: list[str] = []
    fragment_paths: list[str] = []
    env_files: list[str] = []
    status: str | None = None
    main_pid: str | None = None

    for line in lines:
        line_s = line.strip()
        if line_s.startswith("ExecStart="):
            exec_paths.extend(_extract_abs_paths_from_line(line_s))
        elif line_s.startswith("FragmentPath="):
            fragment_paths.extend(_extract_abs_paths_from_line(line_s))
        elif line_s.startswith("EnvironmentFile="):
            env_files.extend(_extract_abs_paths_from_line(line_s))
        elif line_s.startswith("Active:"):
            status = line_s
        elif "Main PID" in line_s:
            main_pid = line_s

    return {
        "unit": unit,
        "exec_paths": list(dict.fromkeys(exec_paths)),
        "fragment_paths": list(dict.fromkeys(fragment_paths)),
        "env_files": list(dict.fromkeys(env_files)),
        "status": status,
        "main_pid": main_pid,
    }


def parse_txt_file(path: Path) -> tuple[list[tuple[str, str, str]], list[tuple[str, str, str]], list[tuple[str, str, str, str]], list[str]]:
    """Parse a host TXT file.

    Returns:
        paths: list of (host, unit, kind, path)
        dirs: list of (host, kind, dir)
        services: list of (host, unit, status, main_pid)
        errors: list of free-form error lines

    Heuristics:
      - host is inferred from filename and optional markers like '==== host ===='.
      - unit is inferred by '--- unit ---' markers OR systemctl status '● unit' lines.
    """
    errors: list[str] = []
    host = _guess_host_from_filename(path.name)

    try:
        raw_lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    except OSError as e:
        return [], [], [], [f"read_error: {path}: {e}"]

    # Optional host marker inside file
    for line in raw_lines[:10]:
        m = re.match(r"====\\s*(.*?)\\s*====", line.strip())
        if m:
            host = m.group(1)
            break

    # Detect unit blocks
    unit_lines: dict[str, list[str]] = {}
    current_unit: str | None = None

    def start_unit(u: str) -> None:
        nonlocal current_unit
        current_unit = u.strip()
        unit_lines.setdefault(current_unit, [])

    for line in raw_lines:
        strip = line.strip()

        # Marker style from scripts: --- UNIT ---
        if re.match(r"^---\\s+.+?\\s+---$", strip):
            start_unit(strip.strip("- "))
            continue

        # systemctl status output starts with bullet
        if strip.startswith("●"):
            # '● nginx.service - ...'
            unit_candidate = strip.lstrip("● ").split()[0]
            if unit_candidate.endswith(".service"):
                start_unit(unit_candidate)
            else:
                # Could be unit without .service
                start_unit(unit_candidate)
            continue

        # If no current unit yet, we still collect paths below.
        if current_unit:
            unit_lines[current_unit].append(line)

    all_paths: list[tuple[str, str, str, str]] = []
    all_dirs_set: set[tuple[str, str, str]] = set()
    services: list[tuple[str, str, str, str]] = []

    # Scan each detected unit block
    for unit, lines in unit_lines.items():
        info = parse_service_block(unit, lines)

        # service record
        services.append((host, unit, info.get("status") or "", info.get("main_pid") or ""))

        # exec paths -> binaries (and app directories)
        for exec_path in info["exec_paths"]:
            all_paths.append((host, unit, "exec", exec_path))
            # app directory: parent of binary (best effort)
            parent = str(Path(exec_path).parent)
            all_dirs_set.add((host, "app_dir", parent))

        # fragment path is usually /etc/systemd/system/.. or /usr/lib/systemd/system/..
        for frag_path in info["fragment_paths"]:
            all_paths.append((host, unit, "unit_fragment", frag_path))

        for env_file in info["env_files"]:
            all_paths.append((host, unit, "env_file", env_file))

        # General absolute paths found inside the block
        for l in lines:
            for p in _extract_abs_paths_from_line(l):
                k = classify_path(p)
                all_paths.append((host, unit, k, p))
                if k in {"log", "config", "certificate", "data"}:
                    all_dirs_set.add((host, f"{k}_dir", str(Path(p).parent)))

    return all_paths, list(all_dirs_set), services, errors


def write_xlsx(out_path: Path, paths: list[tuple[str, str, str, str]], dirs: list[tuple[str, str, str]], services: list[tuple[str, str, str, str]], errors: list[str]) -> None:
    wb = Workbook()

    ws_paths = wb.active
    ws_paths.title = "Paths"
    ws_paths.append(["Host", "Unit", "Kind", "Path"])
    for host, unit, kind, p in paths:
        ws_paths.append([host, unit, kind, p])

    ws_dirs = wb.create_sheet("Dirs")
    ws_dirs.append(["Host", "Kind", "Dir"])
    for host, kind, d in dirs:
        ws_dirs.append([host, kind, d])

    ws_services = wb.create_sheet("Services")
    ws_services.append(["Host", "Unit", "Active/Status", "Main PID"])
    for host, unit, status, pid in services:
        ws_services.append([host, unit, status, pid])

    ws_errors = wb.create_sheet("Errors")
    ws_errors.append(["Error"])
    for e in errors:
        ws_errors.append([e])

    # basic filter on Paths
    ws_paths.auto_filter.ref = ws_paths.dimensions

    wb.save(out_path)


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("Usage: python manual_txt_to_xlsx.py <inputs...> [output.xlsx]", file=sys.stderr)
        return 2

    *inputs, maybe_out = argv
    out = Path(maybe_out) if maybe_out.lower().endswith(".xlsx") else Path("manual_inventory_report.xlsx")
    if out.name == "manual_inventory_report.xlsx":
        inputs.append(maybe_out)

    txt_files: list[Path] = []
    for arg in inputs:
        p = Path(arg)
        if p.is_dir():
            txt_files.extend(Path(x) for x in glob(str(p / "*.txt")))
        else:
            txt_files.append(p)

    if not txt_files:
        print("No input txt files found", file=sys.stderr)
        return 1

    all_paths: list[tuple[str, str, str, str]] = []
    all_dirs: list[tuple[str, str, str]] = []
    all_services: list[tuple[str, str, str, str]] = []
    all_errors: list[str] = []

    for f in txt_files:
        paths, dirs, services, errors = parse_txt_file(f)
        all_paths.extend(paths)
        all_dirs.extend(dirs)
        all_services.extend(services)
        all_errors.extend(errors)

    # de-duplicate across host/unit
    def dedupe(items: list[tuple]) -> list[tuple]:
        seen = set()
        res = []
        for it in items:
            if it in seen:
                continue
            seen.add(it)
            res.append(it)
        return res

    write_xlsx(out, dedupe(all_paths), dedupe(all_dirs), dedupe(all_services), all_errors)
    print(f"[OK] wrote {out} from {len(txt_files)} txt file(s)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
